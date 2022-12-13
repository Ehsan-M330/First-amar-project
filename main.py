from ContinuousDataController import ContinuousData
from Calculation import Cal
import openpyxl
import matplotlib.pyplot as plt
import math

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("data.xlsx")

# Define variable to read sheet
dataframe1 = dataframe.active

# Iterate the loop to read the cell values

if __name__ == "__main__":
    data = []
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(3, 3):
            if col[row].value is not None and col[row].value != "Undergrad Grade":
                data.append(col[row].value)

    quantileNumber = int(input('please input the quantile number: '))

    fig = plt.figure(figsize=(10, 7))
    plt.boxplot(data)
    plt.show()
    average = Cal.average(data)
    variance = Cal.variance(data)
    secondData = ContinuousData(data)
    secondData.printTheTable()
    mode = secondData.calculateModeC()
    quantile = secondData.calculateQuantileC(quantileNumber)
    median = secondData.calculateMedianC()
    print('mode is: ', f'{mode:.2f}')
    print('quantile is: ', f'{quantile:.2f}')
    print('average is: ', f'{average:.2f}')
    print('variance is: ', f'{variance:.2f}')
    print('median is: ', f'{median:.2f}')
